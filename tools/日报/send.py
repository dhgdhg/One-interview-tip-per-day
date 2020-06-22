import openpyxl
import re
import email
import smtplib
import os
import json

from email import encoders
from openpyxl.writer.excel import save_virtual_workbook
from email.mime.application import MIMEApplication
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import date


today = date.today()

def get_content(table):
    name_and_job_number = table.cell(3, 2).value
    day = table.cell(3, 4).value
    group = table.cell(3, 3).value.replace('\n', '</br>')
    progress = table.cell(7, 2).value.replace('\n', '</br>')
    target = table.cell(5, 2).value.replace('\n', '</br>')
    today_process = table.cell(9, 2).value.replace('\n', '</br>')
    risk = table.cell(11, 2).value.replace('\n', '</br>')
    plan = table.cell(13, 2).value.replace('\n', '</br>')

    content = '''
    <html>
    <head>
    <meta http-equiv="Content-Type" content="text/html; charset=utf-8">
    </head>
    <body>
    <div style="color: rgb(33, 33, 33); background-color: rgb(255, 255, 255); text-align: left;" dir="auto">
    <br>
    </div>
    <div id="id-fb9792a3-4bd3-4815-aed4-1d6956e14367" class="ms-outlook-mobile-reference-message" dir="auto" style="text-align: left;">
    <meta content="text/html; charset=utf-8">
    <style>
    <!--
    font
        {{line-height:1.6}}
    ul, ol
        {{padding-left:20px;
        list-style-position:inside}}
    -->
    </style>
    <div style="font-family:微软雅黑,Verdana,&quot;Microsoft Yahei&quot;,SimSun,sans-serif; font-size:14px; line-height:1.6">
    <div></div>
    <style>
    <!--
    font
        {{line-height:1.6}}
    ul, ol
        {{padding-left:20px;
        list-style-position:inside}}
    -->
    </style>
    <div style="font-family:微软雅黑,Verdana,&quot;Microsoft Yahei&quot;,SimSun,sans-serif; font-size:14px; line-height:1.6">
    <div></div>
    <div dir="auto" style="color:rgb(33,33,33); background-color:rgb(255,255,255); text-align:left">
    <table border="0" cellpadding="0" cellspacing="0" width="951" style="border-collapse:collapse; width:714pt">
    <colgroup><col width="72" style="width:54pt"><col width="293" span="3" style="width:220pt"></colgroup>
    <tbody>
    <tr height="19" class="firstRow" style="height:14.4pt">
    <td height="19" width="72" style="height:14.4pt; width:54pt; padding-top:1px; padding-right:1px; padding-left:1px; color:windowtext; font-size:11pt; font-family:宋体; vertical-align:middle; border:none; white-space:nowrap">
    </td>
    <td width="293" style="width:220pt; padding-top:1px; padding-right:1px; padding-left:1px; color:windowtext; font-size:11pt; font-family:宋体; vertical-align:middle; border:none; white-space:nowrap">
    </td>
    <td width="293" style="width:220pt; padding-top:1px; padding-right:1px; padding-left:1px; color:windowtext; font-size:11pt; font-family:宋体; vertical-align:middle; border:none; white-space:nowrap">
    </td>
    <td width="293" style="width:220pt; padding-top:1px; padding-right:1px; padding-left:1px; color:windowtext; font-size:11pt; font-family:宋体; vertical-align:middle; border:none; white-space:nowrap">
    </td>
    </tr>
    <tr height="58" style="height:44.25pt">
    <td height="58" style="height:44.25pt; padding-top:1px; padding-right:1px; padding-left:1px; color:windowtext; font-size:11pt; font-family:宋体; vertical-align:middle; border:none; white-space:nowrap">
    </td>
    <td colspan="3" class="xl72" style="border-width:2pt 1pt 1pt 2pt; border-style:double solid solid double; border-color:windowtext; padding-top:1px; padding-right:1px; padding-left:1px; color:windowtext; font-size:20pt; font-family:微软雅黑,sans-serif; vertical-align:middle; white-space:nowrap; text-align:center; background-image:initial; background-attachment:initial; background-color:rgb(190,215,238); background-size:initial; background-origin:initial; background-clip:initial; background-position:initial; background-repeat:initial">
    openEuler开发项目组工作日报</td>
    </tr>
    <tr height="40" style="height:30.0pt">
    <td height="40" style="height:30pt; padding-top:1px; padding-right:1px; padding-left:1px; color:windowtext; font-size:11pt; font-family:宋体; vertical-align:middle; border:none; white-space:nowrap">
    </td>
    <td class="xl63" style="padding-top:1px; padding-right:1px; padding-left:1px; color:windowtext; font-size:12pt; font-family:微软雅黑,sans-serif; vertical-align:middle; border-style:none none solid double; border-top-width:initial; border-right-width:initial; border-bottom-width:1pt; border-left-width:2pt; border-top-color:initial; border-right-color:initial; border-bottom-color:windowtext; border-left-color:windowtext; white-space:nowrap">
    {name_and_job_number}</td>
    <td class="xl64" style="padding-top:1px; padding-right:1px; padding-left:1px; color:windowtext; font-size:12pt; font-family:微软雅黑,sans-serif; vertical-align:middle; border-style:none none solid solid; border-top-width:initial; border-right-width:initial; border-bottom-width:1pt; border-left-width:1pt; border-top-color:initial; border-right-color:initial; border-bottom-color:windowtext; border-left-color:windowtext; white-space:nowrap">
    {group}</td>
    <td class="xl65" style="padding-top:1px; padding-right:1px; padding-left:1px; color:windowtext; font-size:12pt; font-family:微软雅黑,sans-serif; vertical-align:middle; border-style:none double solid solid; border-top-width:initial; border-right-width:2pt; border-bottom-width:1pt; border-left-width:1pt; border-top-color:initial; border-right-color:windowtext; border-bottom-color:windowtext; border-left-color:windowtext; white-space:nowrap">
    {day}</td>
    </tr>
    <tr height="23" style="height:17.4pt">
    <td height="23" style="height:17.4pt; padding-top:1px; padding-right:1px; padding-left:1px; color:windowtext; font-size:11pt; font-family:宋体; vertical-align:middle; border:none; white-space:nowrap">
    </td>
    <td colspan="3" class="xl66" style="border-right-width:initial; border-style:solid none solid double; border-right-color:initial; padding-top:1px; padding-right:1px; padding-left:1px; color:windowtext; font-size:12pt; font-weight:700; font-family:微软雅黑,sans-serif; vertical-align:middle; border-top-width:1pt; border-bottom-width:1pt; border-left-width:2pt; border-top-color:windowtext; border-bottom-color:windowtext; border-left-color:windowtext; white-space:nowrap; background-image:initial; background-attachment:initial; background-color:rgb(190,215,238); background-size:initial; background-origin:initial; background-clip:initial; background-position:initial; background-repeat:initial">
    工作/学习目标</td>
    </tr>
    <tr height="106" style="height:80.1pt">
    <td height="106" style="height:80.1pt; padding-top:1px; padding-right:1px; padding-left:1px; color:windowtext; font-size:11pt; font-family:宋体; vertical-align:middle; border:none; white-space:nowrap">
    </td>
    <td colspan="3" class="xl69" width="879" style="border-right-width:initial; border-style:solid none solid double; border-right-color:initial; width:660pt; padding-top:1px; padding-right:1px; padding-left:1px; color:windowtext; font-size:12pt; font-family:微软雅黑,sans-serif; vertical-align:top; border-top-width:1pt; border-bottom-width:1pt; border-left-width:2pt; border-top-color:windowtext; border-bottom-color:windowtext; border-left-color:windowtext">
    {target}
    </td>
    </tr>
    <tr height="23" style="height:17.4pt">
    <td height="23" style="height:17.4pt; padding-top:1px; padding-right:1px; padding-left:1px; color:windowtext; font-size:11pt; font-family:宋体; vertical-align:middle; border:none; white-space:nowrap">
    </td>
    <td colspan="3" class="xl66" style="border-right-width:initial; border-style:solid none solid double; border-right-color:initial; padding-top:1px; padding-right:1px; padding-left:1px; color:windowtext; font-size:12pt; font-weight:700; font-family:微软雅黑,sans-serif; vertical-align:middle; border-top-width:1pt; border-bottom-width:1pt; border-left-width:2pt; border-top-color:windowtext; border-bottom-color:windowtext; border-left-color:windowtext; white-space:nowrap; background-image:initial; background-attachment:initial; background-color:rgb(190,215,238); background-size:initial; background-origin:initial; background-clip:initial; background-position:initial; background-repeat:initial">
    总体进展</td>
    </tr>
    <tr height="133" style="height:99.9pt">
    <td height="133" style="height:99.9pt; padding-top:1px; padding-right:1px; padding-left:1px; color:windowtext; font-size:11pt; font-family:宋体; vertical-align:middle; border:none; white-space:nowrap">
    </td>
    <td colspan="3" class="xl69" width="879" style="border-right-width:initial; border-style:solid none solid double; border-right-color:initial; width:660pt; padding-top:1px; padding-right:1px; padding-left:1px; color:windowtext; font-size:12pt; font-family:微软雅黑,sans-serif; vertical-align:top; border-top-width:1pt; border-bottom-width:1pt; border-left-width:2pt; border-top-color:windowtext; border-bottom-color:windowtext; border-left-color:windowtext">
    {progress}
    </td>
    </tr>
    <tr height="23" style="height:17.4pt">
    <td height="23" style="height:17.4pt; padding-top:1px; padding-right:1px; padding-left:1px; color:windowtext; font-size:11pt; font-family:宋体; vertical-align:middle; border:none; white-space:nowrap">
    </td>
    <td colspan="3" class="xl66" style="border-right-width:initial; border-style:solid none solid double; border-right-color:initial; padding-top:1px; padding-right:1px; padding-left:1px; color:windowtext; font-size:12pt; font-weight:700; font-family:微软雅黑,sans-serif; vertical-align:middle; border-top-width:1pt; border-bottom-width:1pt; border-left-width:2pt; border-top-color:windowtext; border-bottom-color:windowtext; border-left-color:windowtext; white-space:nowrap; background-image:initial; background-attachment:initial; background-color:rgb(190,215,238); background-size:initial; background-origin:initial; background-clip:initial; background-position:initial; background-repeat:initial">
    今日进展</td>
    </tr>
    <tr height="160" style="height:120.0pt">
    <td height="160" style="height:120pt; padding-top:1px; padding-right:1px; padding-left:1px; color:windowtext; font-size:11pt; font-family:宋体; vertical-align:middle; border:none; white-space:nowrap">
    </td>
    <td colspan="3" class="xl69" width="879" style="border-right-width:initial; border-style:solid none solid double; border-right-color:initial; width:660pt; padding-top:1px; padding-right:1px; padding-left:1px; color:windowtext; font-size:12pt; font-family:微软雅黑,sans-serif; vertical-align:top; border-top-width:1pt; border-bottom-width:1pt; border-left-width:2pt; border-top-color:windowtext; border-bottom-color:windowtext; border-left-color:windowtext">
    {today_process} </td>
    </tr>
    <tr height="23" style="height:17.4pt">
    <td height="23" style="height:17.4pt; padding-top:1px; padding-right:1px; padding-left:1px; color:windowtext; font-size:11pt; font-family:宋体; vertical-align:middle; border:none; white-space:nowrap">
    </td>
    <td colspan="3" class="xl66" style="border-right-width:initial; border-style:solid none solid double; border-right-color:initial; padding-top:1px; padding-right:1px; padding-left:1px; color:windowtext; font-size:12pt; font-weight:700; font-family:微软雅黑,sans-serif; vertical-align:middle; border-top-width:1pt; border-bottom-width:1pt; border-left-width:2pt; border-top-color:windowtext; border-bottom-color:windowtext; border-left-color:windowtext; white-space:nowrap; background-image:initial; background-attachment:initial; background-color:rgb(190,215,238); background-size:initial; background-origin:initial; background-clip:initial; background-position:initial; background-repeat:initial">
    风险及问题</td>
    </tr>
    <tr height="106" style="height:80.1pt">
    <td height="106" style="height:80.1pt; padding-top:1px; padding-right:1px; padding-left:1px; color:windowtext; font-size:11pt; font-family:宋体; vertical-align:middle; border:none; white-space:nowrap">
    </td>
    <td colspan="3" class="xl69" width="879" style="border-right-width:initial; border-style:solid none solid double; border-right-color:initial; width:660pt; padding-top:1px; padding-right:1px; padding-left:1px; color:windowtext; font-size:12pt; font-family:微软雅黑,sans-serif; vertical-align:top; border-top-width:1pt; border-bottom-width:1pt; border-left-width:2pt; border-top-color:windowtext; border-bottom-color:windowtext; border-left-color:windowtext">
    {risk}</td>
    </tr>
    <tr height="23" style="height:17.4pt">
    <td height="23" style="height:17.4pt; padding-top:1px; padding-right:1px; padding-left:1px; color:windowtext; font-size:11pt; font-family:宋体; vertical-align:middle; border:none; white-space:nowrap">
    </td>
    <td colspan="3" class="xl66" style="border-right-width:initial; border-style:solid none solid double; border-right-color:initial; padding-top:1px; padding-right:1px; padding-left:1px; color:windowtext; font-size:12pt; font-weight:700; font-family:微软雅黑,sans-serif; vertical-align:middle; border-top-width:1pt; border-bottom-width:1pt; border-left-width:2pt; border-top-color:windowtext; border-bottom-color:windowtext; border-left-color:windowtext; white-space:nowrap; background-image:initial; background-attachment:initial; background-color:rgb(190,215,238); background-size:initial; background-origin:initial; background-clip:initial; background-position:initial; background-repeat:initial">
    明日计划</td>
    </tr>
    <tr height="106" style="height:80.1pt">
    <td height="106" style="height:80.1pt; padding-top:1px; padding-right:1px; padding-left:1px; color:windowtext; font-size:11pt; font-family:宋体; vertical-align:middle; border:none; white-space:nowrap">
    </td>
    <td colspan="3" class="xl75" width="879" style="border-right-width:initial; border-style:solid none double double; border-right-color:initial; width:660pt; padding-top:1px; padding-right:1px; padding-left:1px; color:windowtext; font-size:12pt; font-family:微软雅黑,sans-serif; vertical-align:top; border-top-width:1pt; border-bottom-width:2pt; border-left-width:2pt; border-top-color:windowtext; border-bottom-color:windowtext; border-left-color:windowtext">
    {plan}</td>
    </tr>
    <tr height="20" style="height:15.0pt">
    <td height="20" style="height:15pt; padding-top:1px; padding-right:1px; padding-left:1px; color:windowtext; font-size:11pt; font-family:宋体; vertical-align:middle; border:none; white-space:nowrap">
    </td>
    <td style="padding-top:1px; padding-right:1px; padding-left:1px; color:windowtext; font-size:11pt; font-family:宋体; vertical-align:middle; border:none; white-space:nowrap">
    </td>
    <td style="padding-top:1px; padding-right:1px; padding-left:1px; color:windowtext; font-size:11pt; font-family:宋体; vertical-align:middle; border:none; white-space:nowrap">
    </td>
    <td style="padding-top:1px; padding-right:1px; padding-left:1px; color:windowtext; font-size:11pt; font-family:宋体; vertical-align:middle; border:none; white-space:nowrap">
    </td>
    </tr>
    </tbody>
    </table>
    </div>
    <style>
    <!--
    font
        {{line-height:1.6}}
    ul, ol
        {{padding-left:20px;
        list-style-position:inside}}
    -->
    </style><style>
    <!--
    #id-0a612769-ea9c-468e-89c5-a3ada09cd41b
        {{line-height:1.5}}
    blockquote
        {{margin-top:0px;
        margin-bottom:0px;
        margin-left:0.5em}}
    #id-0a612769-ea9c-468e-89c5-a3ada09cd41b
        {{font-size:10.5pt;
        font-family:"Microsoft YaHei UI";
        color:rgb(0,0,0);
        line-height:1.5}}
    -->
    </style><style>
    <!--
    -->
    </style></div>
    </div>
    <br>
    </div>
    </body>
    </html>
    '''.format(
        name_and_job_number=name_and_job_number,
        day=day,
        group=group,
        progress=progress,
        target=target,
        today_process=today_process,
        risk=risk,
        plan=plan
    )

    return content


def get_sheet():
    xlsx = openpyxl.load_workbook('template.xlsx')
    sheet = xlsx.active
    sheet.cell(row=3, column=4).value = '汇报日期：{}'.format(today)
    return xlsx


def send_email():
    settings = {}
    with open('settings.json') as f:
        settings = json.loads(f.read())

    sender = settings['sender']
    name = settings['name']
    password = settings['password']
    receiver = settings['receiver']
    copy_to = settings['copy_to']
    subject = settings['subject_template'].format(today=today, name=name)

    msg = MIMEMultipart()
    sheet = get_sheet()
    email_file = MIMEApplication(save_virtual_workbook((sheet)))
    encoders.encode_base64(email_file)
    email_file.set_payload(email_file.get_payload())
    email_file.add_header('Content-Disposition', 'attachment', filename='{}日报{}-{}.xlsx'.format(name, today.month, today.day))
    msg.attach(email_file)

    msg.attach(MIMEText(get_content(sheet.worksheets[1]), _subtype='html', _charset='utf-8'))
    msg['From'] = sender
    msg['To'] = receiver
    msg['Subject'] = subject
    msg['Cc'] = copy_to
    
    smtp = smtplib.SMTP("smtp.live.com", 587)
    smtp.set_debuglevel(1)
    smtp.ehlo()
    smtp.starttls()
    smtp.ehlo()
    smtp.login(sender, password)
    smtp.sendmail(sender, receiver, msg.as_string())
    smtp.quit()

if __name__ == "__main__":
    send_email()
